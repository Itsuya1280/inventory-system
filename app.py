import os
from flask import Flask, redirect, url_for, render_template, request, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, current_user, login_user, logout_user
from datetime import datetime
from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import UserMixin

app = Flask(__name__)
app.config['SECRET_KEY'] = 'inventory-system-secret-key-2024'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///inventory.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['WTF_CSRF_ENABLED'] = False

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login_page'

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False, index=True)
    username = db.Column(db.String(80), nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class ItemGroup(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    display_order = db.Column(db.Integer, nullable=False, default=0)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)

class Stock(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    product_name = db.Column(db.String(100), nullable=False)
    quantity = db.Column(db.Integer, nullable=False, default=0)
    supplier = db.Column(db.String(100))
    group_id = db.Column(db.Integer, db.ForeignKey('item_group.id'))
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow, onupdate=datetime.utcnow)
    deleted_at = db.Column(db.DateTime)
    group = db.relationship('ItemGroup', backref='stocks')

class StockHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    stock_id = db.Column(db.Integer, db.ForeignKey('stock.id'), nullable=False)
    quantity_change = db.Column(db.Integer, nullable=False)
    transaction_type = db.Column(db.String(20), nullable=False)
    reference_id = db.Column(db.Integer)
    notes = db.Column(db.String(200))
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    stock = db.relationship('Stock', backref='history')
    user = db.relationship('User', backref='operations')

class OutboundOrder(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    stock_id = db.Column(db.Integer, db.ForeignKey('stock.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    destination = db.Column(db.String(200), nullable=False)
    status = db.Column(db.String(20), nullable=False, default='pending')
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    warehouse_confirmed_at = db.Column(db.DateTime)
    completed_at = db.Column(db.DateTime)
    stock = db.relationship('Stock', backref='outbound_orders')

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login_page'))

@app.errorhandler(404)
def not_found(error):
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def server_error(error):
    return render_template('errors/500.html'), 500

@app.route('/auth/login', methods=['GET', 'POST'])
def login_page():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        user = User.query.filter_by(email=email).first()
        if user and user.check_password(password):
            login_user(user)
            return redirect(url_for('dashboard'))
        flash('ログイン失敗', 'error')
    
    return render_template('auth/login.html')

@app.route('/auth/logout')
def logout():
    logout_user()
    return redirect(url_for('login_page'))

@app.route('/dashboard')
def dashboard():
    if not current_user.is_authenticated:
        return redirect(url_for('login_page'))
    
    from sqlalchemy import func
    total_items = Stock.query.filter(Stock.deleted_at.is_(None)).count()
    total_quantity = db.session.query(func.sum(Stock.quantity)).filter(Stock.deleted_at.is_(None)).scalar() or 0
    total_groups = ItemGroup.query.count()
    
    return render_template('dashboard/index.html', total_items=total_items, total_quantity=total_quantity, total_groups=total_groups)

@app.route('/inventory')
def inventory_list():
    if not current_user.is_authenticated:
        return redirect(url_for('login_page'))
    
    search = request.args.get('search', '').strip()
    group_filter = request.args.get('group', type=int)
    supplier_filter = request.args.get('supplier', '').strip()
    
    query = Stock.query.filter(Stock.deleted_at.is_(None))
    if search:
        query = query.filter(Stock.product_name.ilike(f'%{search}%'))
    if group_filter:
        query = query.filter(Stock.group_id == group_filter)
    if supplier_filter:
        query = query.filter(Stock.supplier.ilike(f'%{supplier_filter}%'))
    
    stocks = query.all()
    groups = ItemGroup.query.all()
    
    # 仕入先の一覧を取得（ユニーク）
    suppliers = db.session.query(Stock.supplier).filter(
        Stock.deleted_at.is_(None),
        Stock.supplier.isnot(None)
    ).distinct().order_by(Stock.supplier.asc()).all()
    suppliers = [s[0] for s in suppliers]
    
    return render_template('inventory/index.html', stocks=stocks, groups=groups, suppliers=suppliers, search=search, group_filter=group_filter, supplier_filter=supplier_filter)

@app.route('/inventory/<int:stock_id>/edit', methods=['GET', 'POST'])
def inventory_edit(stock_id):
    if not current_user.is_authenticated:
        return redirect(url_for('login_page'))
    
    stock = Stock.query.get_or_404(stock_id)
    
    if request.method == 'POST':
        product_name = request.form.get('product_name', '').strip()
        quantity = request.form.get('quantity', type=int)
        group_id = request.form.get('group_id', type=int)
        supplier = request.form.get('supplier', '').strip()
        
        if not product_name:
            flash('商品名を入力してください', 'error')
        elif quantity is None or quantity < 0:
            flash('数量を正しく入力してください', 'error')
        else:
            try:
                stock.product_name = product_name
                stock.quantity = quantity
                stock.group_id = group_id
                stock.supplier = supplier
                stock.updated_at = datetime.utcnow()
                db.session.commit()
                flash('在庫情報を更新しました', 'success')
                return redirect(url_for('inventory_list'))
            except Exception as e:
                db.session.rollback()
                flash(f'エラー: {str(e)}', 'error')
    
    groups = ItemGroup.query.all()
    return render_template('inventory/edit.html', stock=stock, groups=groups)

@app.route('/inventory/<int:stock_id>/delete', methods=['POST'])
def inventory_delete(stock_id):
    if not current_user.is_authenticated:
        return redirect(url_for('login_page'))
    
    try:
        stock = Stock.query.get_or_404(stock_id)
        stock.deleted_at = datetime.utcnow()
        db.session.commit()
        flash('在庫を削除しました', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'エラー: {str(e)}', 'error')
    
    return redirect(url_for('inventory_list'))

@app.route('/inbound')
def inbound_index():
    if not current_user.is_authenticated:
        return redirect(url_for('login_page'))
    return render_template('inbound/index.html')

@app.route('/inbound/new', methods=['GET', 'POST'])
def inbound_new():
    if not current_user.is_authenticated:
        return redirect(url_for('login_page'))
    
    if request.method == 'POST':
        group_id = request.form.get('group_id', type=int)
        product_name = request.form.get('product_name', '').strip()
        quantity = request.form.get('quantity', type=int)
        supplier = request.form.get('supplier', '').strip()
        
        if not group_id:
            flash('グループを選択してください', 'error')
        elif not product_name:
            flash('枝番を入力してください', 'error')
        elif not quantity or quantity <= 0:
            flash('数量を正しく入力してください', 'error')
        elif not supplier:
            flash('仕入先を入力してください', 'error')
        else:
            try:
                group = ItemGroup.query.get(group_id)
                if not group:
                    flash('グループが見つかりません', 'error')
                    return redirect(url_for('inbound_new'))
                
                existing_stock = Stock.query.filter_by(group_id=group_id, product_name=product_name).filter(Stock.deleted_at.is_(None)).first()
                
                if existing_stock:
                    existing_stock.quantity += quantity
                    existing_stock.supplier = supplier
                    existing_stock.updated_at = datetime.utcnow()
                else:
                    stock = Stock(product_name=product_name, quantity=quantity, group_id=group_id, supplier=supplier)
                    db.session.add(stock)
                    db.session.flush()
                    existing_stock = stock
                
                history = StockHistory(stock_id=existing_stock.id, quantity_change=quantity, transaction_type='inbound', notes=f'入庫: {supplier}', user_id=current_user.id)
                db.session.add(history)
                db.session.commit()
                
                flash(f'{product_name} を {quantity}個 入庫しました', 'success')
                return redirect(url_for('inbound_index'))
            except Exception as e:
                db.session.rollback()
                flash(f'エラー: {str(e)}', 'error')
    
    groups = ItemGroup.query.order_by(ItemGroup.display_order.asc(), ItemGroup.created_at.desc()).all()
    return render_template('inbound/new.html', groups=groups)

@app.route('/inbound/api/stocks/<group_id>')
def inbound_get_stocks(group_id):
    try:
        group_id = int(group_id)
        stocks = Stock.query.filter(Stock.group_id == group_id, Stock.deleted_at.is_(None)).all()
        return jsonify([{'id': s.id, 'product_name': s.product_name, 'quantity': s.quantity} for s in stocks])
    except:
        return jsonify([])

@app.route('/outbound')
def outbound_index():
    if not current_user.is_authenticated:
        return redirect(url_for('login_page'))
    pending_orders = OutboundOrder.query.filter_by(status='pending').all()
    return render_template('outbound/index.html', pending_orders=pending_orders)

@app.route('/outbound/new', methods=['GET', 'POST'])
def outbound_new():
    if not current_user.is_authenticated:
        return redirect(url_for('login_page'))
    
    if request.method == 'POST':
        group_id = request.form.get('group_id', type=int)
        stock_id = request.form.get('stock_id', type=int)
        quantity = request.form.get('quantity', type=int)
        destination = request.form.get('destination', '').strip()
        
        if not group_id:
            flash('グループを選択してください', 'error')
        elif not stock_id:
            flash('枝番を選択してください', 'error')
        elif not quantity or quantity <= 0:
            flash('数量を正しく入力してください', 'error')
        elif not destination:
            flash('出荷先を入力してください', 'error')
        else:
            try:
                stock = Stock.query.get(stock_id)
                if not stock or stock.quantity < quantity:
                    flash('在庫が不足しています', 'error')
                    return redirect(url_for('outbound_new'))
                
                order = OutboundOrder(stock_id=stock_id, quantity=quantity, destination=destination, status='pending')
                db.session.add(order)
                db.session.flush()
                
                stock.quantity -= quantity
                stock.updated_at = datetime.utcnow()
                
                history = StockHistory(stock_id=stock_id, quantity_change=-quantity, transaction_type='outbound', reference_id=order.id, notes=f'出庫: {destination}', user_id=current_user.id)
                db.session.add(history)
                db.session.commit()
                
                flash(f'{stock.product_name} を {quantity}個 出庫予定にしました', 'success')
                return redirect(url_for('outbound_index'))
            except Exception as e:
                db.session.rollback()
                flash(f'エラー: {str(e)}', 'error')
    
    groups = ItemGroup.query.order_by(ItemGroup.display_order.asc(), ItemGroup.created_at.desc()).all()
    return render_template('outbound/new.html', groups=groups)

@app.route('/outbound/api/stocks/<group_id>')
def outbound_get_stocks(group_id):
    try:
        group_id = int(group_id)
        stocks = Stock.query.filter(Stock.group_id == group_id, Stock.deleted_at.is_(None), Stock.quantity > 0).all()
        return jsonify([{'id': s.id, 'product_name': s.product_name, 'quantity': s.quantity} for s in stocks])
    except:
        return jsonify([])

@app.route('/outbound/<int:order_id>/cancel', methods=['POST'])
def outbound_cancel(order_id):
    if not current_user.is_authenticated:
        return redirect(url_for('login_page'))
    
    try:
        order = OutboundOrder.query.get(order_id)
        if not order or order.status != 'pending':
            flash('キャンセルできません', 'error')
            return redirect(url_for('outbound_index'))
        
        stock = Stock.query.get(order.stock_id)
        stock.quantity += order.quantity
        stock.updated_at = datetime.utcnow()
        
        db.session.delete(order)
        db.session.commit()
        
        flash('出庫予定をキャンセルしました', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'エラー: {str(e)}', 'error')
    
    return redirect(url_for('outbound_index'))

@app.route('/warehouse')
def warehouse_index():
    if not current_user.is_authenticated:
        return redirect(url_for('login_page'))
    
    pending_orders = OutboundOrder.query.filter_by(status='pending').order_by(OutboundOrder.created_at).all()
    confirmed_orders = OutboundOrder.query.filter_by(status='warehouse_confirmed').order_by(OutboundOrder.warehouse_confirmed_at.desc()).all()
    completed_orders = OutboundOrder.query.filter_by(status='completed').order_by(OutboundOrder.completed_at.desc()).all()
    
    return render_template('warehouse/index.html', pending_orders=pending_orders, confirmed_orders=confirmed_orders, completed_orders=completed_orders)

@app.route('/warehouse/<int:order_id>/confirm', methods=['POST'])
def warehouse_confirm(order_id):
    if not current_user.is_authenticated:
        return jsonify({'success': False, 'message': 'ログインしてください'}), 401
    
    try:
        order = OutboundOrder.query.get_or_404(order_id)
        if order.status != 'pending':
            return jsonify({'success': False, 'message': 'このオーダーは確認済みです'}), 400
        
        order.status = 'warehouse_confirmed'
        order.warehouse_confirmed_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({'success': True, 'message': '倉庫確認を完了しました'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'エラー: {str(e)}'}), 500

@app.route('/warehouse/<int:order_id>/complete', methods=['POST'])
def warehouse_complete(order_id):
    if not current_user.is_authenticated:
        return jsonify({'success': False, 'message': 'ログインしてください'}), 401
    
    try:
        order = OutboundOrder.query.get_or_404(order_id)
        if order.status != 'warehouse_confirmed':
            return jsonify({'success': False, 'message': 'この操作はできません'}), 400
        
        order.status = 'completed'
        order.completed_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({'success': True, 'message': '出庫完了にしました'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'エラー: {str(e)}'}), 500

@app.route('/warehouse/<int:order_id>/revert', methods=['POST'])
def warehouse_revert(order_id):
    if not current_user.is_authenticated:
        return jsonify({'success': False, 'message': 'ログインしてください'}), 401
    
    try:
        order = OutboundOrder.query.get_or_404(order_id)
        
        if order.status == 'completed':
            order.status = 'warehouse_confirmed'
            order.completed_at = None
        elif order.status == 'warehouse_confirmed':
            order.status = 'pending'
            order.warehouse_confirmed_at = None
        else:
            return jsonify({'success': False, 'message': 'この操作はできません'}), 400
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'ステータスを戻しました'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'エラー: {str(e)}'}), 500

@app.route('/history')
def history_list():
    if not current_user.is_authenticated:
        return redirect(url_for('login_page'))
    
    transaction_type = request.args.get('type', '').strip()
    search_product = request.args.get('search_product', '').strip()
    group_filter = request.args.get('group', type=int)
    user_filter = request.args.get('user', type=int)
    destination_filter = request.args.get('destination', '').strip()
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    
    query = StockHistory.query.order_by(StockHistory.created_at.desc())
    
    if transaction_type in ['inbound', 'outbound', 'adjustment']:
        query = query.filter_by(transaction_type=transaction_type)
    
    if search_product:
        query = query.join(Stock).filter(Stock.product_name.ilike(f'%{search_product}%'))
    
    if group_filter:
        query = query.join(Stock).filter(Stock.group_id == group_filter)
    
    if user_filter:
        query = query.filter_by(user_id=user_filter)
    
    if destination_filter:
        query = query.filter(StockHistory.notes.ilike(f'%{destination_filter}%'))
    
    if start_date:
        try:
            start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(StockHistory.created_at >= start_datetime)
        except:
            pass
    
    if end_date:
        try:
            end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
            end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
            query = query.filter(StockHistory.created_at <= end_datetime)
        except:
            pass
    
    history = query.all()
    
    # フィルター用のデータ取得
    groups = ItemGroup.query.order_by(ItemGroup.display_order.asc()).all()
    users = User.query.order_by(User.username.asc()).all()
    
    return render_template('history/index.html', 
                         history=history, 
                         transaction_type=transaction_type, 
                         search_product=search_product,
                         group_filter=group_filter,
                         user_filter=user_filter,
                         destination_filter=destination_filter,
                         start_date=start_date,
                         end_date=end_date,
                         groups=groups,
                         users=users)

@app.route('/item_master')
def item_master_index():
    if not current_user.is_authenticated:
        return redirect(url_for('login_page'))
    
    groups = ItemGroup.query.order_by(ItemGroup.display_order.asc(), ItemGroup.created_at.desc()).all()
    
    group_data = []
    for idx, group in enumerate(groups):
        count = Stock.query.filter(Stock.group_id == group.id, Stock.deleted_at.is_(None)).count()
        group_data.append({'group': group, 'count': count, 'order': idx})
    
    return render_template('item_master/index.html', group_data=group_data)

@app.route('/item_master/api/reorder', methods=['POST'])
def item_master_reorder():
    if not current_user.is_authenticated:
        return jsonify({'success': False, 'message': 'ログインしてください'}), 401
    
    try:
        data = request.get_json()
        orders = data.get('orders', [])
        
        for idx, group_id in enumerate(orders):
            group = ItemGroup.query.get(int(group_id))
            if group:
                group.display_order = idx
        
        db.session.commit()
        return jsonify({'success': True, 'message': '並べ替えを保存しました'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'エラー: {str(e)}'}), 500

@app.route('/item_master/group/new', methods=['GET', 'POST'])
def item_master_new_group():
    if not current_user.is_authenticated:
        return redirect(url_for('login_page'))
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        
        if not name:
            flash('グループ名を入力してください', 'error')
            return render_template('item_master/new_group.html')
        
        try:
            existing = ItemGroup.query.filter_by(name=name).first()
            if existing:
                flash(f'グループ「{name}」は既に存在します', 'error')
                return render_template('item_master/new_group.html')
            
            max_order = db.session.query(db.func.max(ItemGroup.display_order)).scalar() or 0
            group = ItemGroup(name=name, display_order=max_order + 1)
            db.session.add(group)
            db.session.commit()
            flash(f'グループ「{name}」を作成しました', 'success')
            return redirect(url_for('item_master_index'))
        except Exception as e:
            db.session.rollback()
            flash(f'エラー: {str(e)}', 'error')
            return render_template('item_master/new_group.html')
    
    return render_template('item_master/new_group.html')

@app.route('/user_management')
def user_management():
    if not current_user.is_authenticated:
        return redirect(url_for('login_page'))
    
    if current_user.email != 'admin@example.com':
        flash('管理者のみアクセス可能です', 'error')
        return redirect(url_for('dashboard'))
    
    sort_by = request.args.get('sort', 'created_at')
    sort_order = request.args.get('order', 'desc')
    
    if sort_order == 'asc':
        if sort_by == 'email':
            users = User.query.order_by(User.email.asc()).all()
        elif sort_by == 'username':
            users = User.query.order_by(User.username.asc()).all()
        else:
            users = User.query.order_by(User.created_at.asc()).all()
    else:
        if sort_by == 'email':
            users = User.query.order_by(User.email.desc()).all()
        elif sort_by == 'username':
            users = User.query.order_by(User.username.desc()).all()
        else:
            users = User.query.order_by(User.created_at.desc()).all()
    
    return render_template('user_management/index.html', users=users, sort_by=sort_by, sort_order=sort_order)

@app.route('/user_management/new', methods=['GET', 'POST'])
def user_management_new():
    if not current_user.is_authenticated:
        return redirect(url_for('login_page'))
    
    if current_user.email != 'admin@example.com':
        flash('管理者のみアクセス可能です', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '').strip()
        
        if not username:
            flash('ユーザー名を入力してください', 'error')
        elif not email:
            flash('メールアドレスを入力してください', 'error')
        elif not password:
            flash('パスワードを入力してください', 'error')
        else:
            try:
                existing = User.query.filter_by(email=email).first()
                if existing:
                    flash(f'メールアドレス「{email}」は既に登録済みです', 'error')
                    return render_template('user_management/new.html')
                
                user = User(username=username, email=email)
                user.set_password(password)
                db.session.add(user)
                db.session.commit()
                flash(f'ユーザー「{username}」を作成しました', 'success')
                return redirect(url_for('user_management'))
            except Exception as e:
                db.session.rollback()
                flash(f'エラー: {str(e)}', 'error')
    
    return render_template('user_management/new.html')

@app.route('/user_management/<int:user_id>/edit', methods=['GET', 'POST'])
def user_management_edit(user_id):
    if not current_user.is_authenticated:
        return redirect(url_for('login_page'))
    
    if current_user.email != 'admin@example.com':
        flash('管理者のみアクセス可能です', 'error')
        return redirect(url_for('dashboard'))
    
    user = User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '').strip()
        
        if not username:
            flash('ユーザー名を入力してください', 'error')
        elif not email:
            flash('メールアドレスを入力してください', 'error')
        else:
            try:
                existing = User.query.filter(User.email == email, User.id != user_id).first()
                if existing:
                    flash(f'メールアドレス「{email}」は既に登録済みです', 'error')
                    return render_template('user_management/edit.html', user=user)
                
                user.username = username
                user.email = email
                
                if password:
                    user.set_password(password)
                
                db.session.commit()
                flash(f'ユーザー「{username}」を更新しました', 'success')
                return redirect(url_for('user_management'))
            except Exception as e:
                db.session.rollback()
                flash(f'エラー: {str(e)}', 'error')
    
    return render_template('user_management/edit.html', user=user)

@app.route('/user_management/<int:user_id>/delete', methods=['POST'])
def user_management_delete(user_id):
    if not current_user.is_authenticated:
        return redirect(url_for('login_page'))
    
    if current_user.email != 'admin@example.com':
        flash('管理者のみアクセス可能です', 'error')
        return redirect(url_for('user_management'))
    
    if user_id == current_user.id:
        flash('自分自身は削除できません', 'error')
        return redirect(url_for('user_management'))
    
    try:
        user = User.query.get_or_404(user_id)
        username = user.username
        db.session.delete(user)
        db.session.commit()
        flash(f'ユーザー「{username}」を削除しました', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'エラー: {str(e)}', 'error')
    
    return redirect(url_for('user_management'))

def init_db():
    with app.app_context():
        db.create_all()
        
        existing_user = User.query.filter_by(email='admin@example.com').first()
        if not existing_user:
            user = User(email='admin@example.com', username='admin')
            user.set_password('Admin@12345')
            db.session.add(user)
            db.session.commit()
            print('✓ テストユーザーを作成しました')
            print('  Email: admin@example.com')
            print('  Password: Admin@12345')
        else:
            print('✓ テストユーザーは既に存在します')

# ========== Excel出力・アップロード ==========

@app.route('/inventory/export')
def inventory_export():
    if not current_user.is_authenticated:
        return redirect(url_for('login_page'))
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        stocks = Stock.query.filter(Stock.deleted_at.is_(None)).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = '在庫一覧'
        
        # ヘッダー行
        headers = ['ID', 'グループ', '商品名（枝番）', '仕入先', '数量']
        ws.append(headers)
        
        # ヘッダーのスタイル
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # データ行
        for stock in stocks:
            ws.append([
                stock.id,
                stock.group.name if stock.group else '-',
                stock.product_name,
                stock.supplier if stock.supplier else '-',
                stock.quantity
            ])
        
        # 列幅調整
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 12
        
        # 枠線
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.border = thin_border
                if cell.row != 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # ダウンロード
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'inventory_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        flash(f'エラー: {str(e)}', 'error')
        return redirect(url_for('inventory_list'))

@app.route('/inventory/import', methods=['GET', 'POST'])
def inventory_import():
    if not current_user.is_authenticated:
        return redirect(url_for('login_page'))
    
    if request.method == 'POST':
        try:
            if 'file' not in request.files:
                flash('ファイルを選択してください', 'error')
                return redirect(url_for('inventory_import'))
            
            file = request.files['file']
            if file.filename == '':
                flash('ファイルを選択してください', 'error')
                return redirect(url_for('inventory_import'))
            
            if not file.filename.endswith('.xlsx'):
                flash('Excelファイル（.xlsx）をアップロードしてください', 'error')
                return redirect(url_for('inventory_import'))
            
            from openpyxl import load_workbook
            from io import BytesIO
            
            wb = load_workbook(BytesIO(file.read()))
            ws = wb.active
            
            updated_count = 0
            error_rows = []
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # 列のデータを取得
                    stock_id = row[0]  # A列
                    quantity = row[4]  # E列（5番目）
                    
                    if stock_id is None or quantity is None:
                        continue
                    
                    # 型変換
                    try:
                        stock_id = int(stock_id)
                    except (ValueError, TypeError):
                        error_rows.append(f'{idx}行目: ID「{stock_id}」は数値で入力してください')
                        continue
                    
                    try:
                        quantity = int(quantity)
                    except (ValueError, TypeError):
                        error_rows.append(f'{idx}行目: 数量「{quantity}」は数値で入力してください（E列を確認）')
                        continue
                    
                    stock = Stock.query.get(stock_id)
                    if not stock:
                        error_rows.append(f'{idx}行目: ID {stock_id} が見つかりません')
                        continue
                    
                    old_quantity = stock.quantity
                    stock.quantity = quantity
                    stock.updated_at = datetime.utcnow()
                    
                    # 差分を履歴に記録
                    quantity_change = quantity - old_quantity
                    if quantity_change != 0:
                        history = StockHistory(
                            stock_id=stock.id,
                            quantity_change=quantity_change,
                            transaction_type='adjustment',
                            notes=f'一括変更: {old_quantity}個 → {quantity}個',
                            user_id=current_user.id
                        )
                        db.session.add(history)
                        updated_count += 1
                
                except Exception as e:
                    error_rows.append(f'{idx}行目: {str(e)}')
            
            db.session.commit()
            
            if error_rows:
                error_msg = '更新完了しましたが、以下の行でエラーが発生しました:\n' + '\n'.join(error_rows[:10])
                if len(error_rows) > 10:
                    error_msg += f'\n... 他 {len(error_rows) - 10} 件'
                flash(error_msg, 'error')
            else:
                flash(f'{updated_count}個の商品を更新しました', 'success')
            
            return redirect(url_for('inventory_list'))
        
        except Exception as e:
            db.session.rollback()
            flash(f'エラー: {str(e)}', 'error')
            return redirect(url_for('inventory_import'))
    
    return render_template('inventory/import.html')

# ========== QRコード機能 ==========

@app.route('/inventory/qr')
def inventory_qr():
    if not current_user.is_authenticated:
        return redirect(url_for('login_page'))
    
    selected_ids = request.args.get('ids', '')
    if not selected_ids:
        flash('商品を選択してください', 'error')
        return redirect(url_for('inventory_list'))
    
    stock_ids = [int(id) for id in selected_ids.split(',') if id.isdigit()]
    stocks = Stock.query.filter(Stock.id.in_(stock_ids)).all()
    
    return render_template('inventory/qr.html', stocks=stocks)

@app.route('/qr/<int:stock_id>')
def qr_detail(stock_id):
    if not current_user.is_authenticated:
        return redirect(url_for('login_page'))
    
    stock = Stock.query.get_or_404(stock_id)
    history = StockHistory.query.filter_by(stock_id=stock_id).order_by(StockHistory.created_at.desc()).limit(20).all()
    
    return render_template('qr/detail.html', stock=stock, history=history)

@app.route('/api/qr/generate/<int:stock_id>')
def api_qr_generate(stock_id):
    if not current_user.is_authenticated:
        return jsonify({'success': False, 'message': 'ログインしてください'}), 401
    
    try:
        import qrcode
        from io import BytesIO
        import base64
        
        stock = Stock.query.get_or_404(stock_id)
        
        qr_url = url_for('qr_detail', stock_id=stock_id, _external=True)
        qr = qrcode.QRCode(version=1, box_size=10, border=2)
        qr.add_data(qr_url)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color='black', back_color='white')
        img_byte_arr = BytesIO()
        img.save(img_byte_arr, format='PNG')
        img_byte_arr.seek(0)
        
        img_base64 = base64.b64encode(img_byte_arr.getvalue()).decode()
        
        return jsonify({
            'success': True,
            'qr_code': f'data:image/png;base64,{img_base64}',
            'stock_id': stock_id,
            'group_name': stock.group.name if stock.group else '-',
            'product_name': stock.product_name
        })
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        return jsonify({'success': False, 'message': str(e), 'traceback': error_msg}), 500

if __name__ == '__main__':
    print('='*50)
    print('  在庫管理システム')
    print('='*50)
    
    init_db()
    
    import os
    env = os.environ.get('FLASK_ENV', 'development')
    
    if env == 'production':
        print('本番環境で実行中...')
        print('Gunicornで起動してください: gunicorn -c gunicorn_config.py app:app')
    else:
        print('')
        print('開発環境で実行中...')
        print('ブラウザを開いてください: http://localhost:5000')
        print('')
        app.run(debug=True, host='0.0.0.0', port=5000, use_reloader=False)
